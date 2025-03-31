import os
import json
import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def extract_neighbors_by_elementid(df):
    """
    content (chunk_with_neighbors) 생성 함수  
    → 이전청크, 현재청크, 다음청크 각각에 라벨을 붙여 결합
    """
    df["elementid"] = df["elementid"].astype(int)
    df_sorted = df.sort_values(by="elementid").reset_index(drop=True)
    # "내용" 열은 이미지설명이 반영된 상태라고 가정
    elementid_to_content = dict(zip(df_sorted["elementid"], df_sorted["내용"]))
    content_list = []
    for eid in df_sorted["elementid"]:
        prev = elementid_to_content.get(eid - 1, "")
        curr = elementid_to_content.get(eid, "")
        next_ = elementid_to_content.get(eid + 1, "")
        parts = [
            "[[[[[[이전청크]", prev,
            "[[[[[[현재청크]", curr,
            "[[[[[[다음청크]", next_
        ]
        combined = "\n\n".join(parts)
        content_list.append(combined)
    return content_list

def extract_page_plus_chunk(df):
    """
    content (page_plus_chunk) 생성 함수  
    → 현재페이지 전체내용과 현재청크에 대해 라벨을 붙여 결합
    """
    df["페이지숫자"] = df["페이지숫자"].astype(str)
    page_to_all_text = df.groupby("페이지숫자")["내용"].apply(lambda x: "\n\n".join(x)).to_dict()
    content_list = []
    for _, row in df.iterrows():
        page = row["페이지숫자"]
        chunk = row["내용"]
        full_page = page_to_all_text.get(page, "")
        parts = [
            "[[[[[[현재페이지 전체내용]", full_page,
            "[[[[[[현재청크]", chunk
        ]
        combined = "\n\n".join(parts)
        content_list.append(combined)
    return content_list

def extract_page_only(df):
    """
    content (page_only) 생성 함수  
    → 현재페이지 전체내용을 라벨과 함께 표시
    """
    df["페이지숫자"] = df["페이지숫자"].astype(str)
    page_to_all_text = df.groupby("페이지숫자")["내용"].apply(lambda x: "\n\n".join(x)).to_dict()
    content_list = []
    for _, row in df.iterrows():
        page = row["페이지숫자"]
        full_page = page_to_all_text.get(page, "")
        combined = "[[[[[[현재페이지 전체내용]\n\n" + full_page
        content_list.append(combined)
    return content_list

def get_neighbor_metadata(df):
    """
    metadata -1 생성 함수 (청크 기반)  
    → 이전청크, 현재청크, 다음청크를 라벨과 함께 결합하고, 
        각 행의 메타데이터를 JSON 객체로 생성  

    JSON 객체 형식:
    {
        "elementid": [int],         # 현재 행의 elementid (리스트)
        "category": str,            # data-category 값
        "filename": str,            # 파일명
        "page": [int],              # 현재 행의 페이지숫자 (리스트)
        "text": str               # 라벨이 포함된 결합 텍스트
    }
    """
    df["elementid"] = df["elementid"].astype(int)
    df["페이지숫자"] = df["페이지숫자"].astype(int)
    df_sorted = df.sort_values(by="elementid").reset_index(drop=True)
    elementid_to_content = dict(zip(df_sorted["elementid"], df_sorted["내용"]))
    metadata = []
    for _, row in df_sorted.iterrows():
        eid = row["elementid"]
        prev = elementid_to_content.get(eid - 1, "")
        curr = elementid_to_content.get(eid, "")
        next_ = elementid_to_content.get(eid + 1, "")
        parts = [
            "[[[[[[이전청크]", prev,
            "[[[[[[현재청크]", curr,
            "[[[[[[다음청크]", next_
        ]
        combined_text = "\n\n".join(parts)
        meta_obj = {
            "elementid": [eid],
            "category": row["data-category"],
            "filename": row["파일명"],
            "page": [row["페이지숫자"]],
            "text": combined_text
        }
        metadata.append(meta_obj)
    return metadata

def get_3page_metadata(df):
    """
    metadata -2 생성 함수 (페이지 기반)  
    → 이전페이지, 현재페이지, 다음페이지 전체 내용을 라벨과 함께 결합하고,
       해당 페이지에 속한 모든 elementid와 페이지 숫자를 리스트로 기록  
       
    JSON 객체 형식:
    {
        "elementid": [int, int, ...],  # 해당 페이지(이전+현재+다음)의 모든 elementid
        "category": str,               # 현재 행의 data-category
        "filename": str,               # 파일명
        "page": [int, int, ...],       # 해당 페이지(이전+현재+다음)의 페이지숫자 리스트
        "text": str                  # 라벨 포함 결합 텍스트
    }
    """
    df["페이지숫자"] = df["페이지숫자"].astype(int)
    # 페이지별 전체 텍스트
    page_to_text = df.groupby("페이지숫자")["내용"].apply(lambda x: "\n\n".join(x)).to_dict()
    metadata = []
    for _, row in df.iterrows():
        current_page = int(row["페이지숫자"])
        prev_text = page_to_text.get(current_page - 1, "")
        current_text = page_to_text.get(current_page, "")
        next_text = page_to_text.get(current_page + 1, "")
        parts = [
            "[[[[[[이전페이지]", prev_text,
            "[[[[[[현재페이지]", current_text,
            "[[[[[[다음페이지]", next_text
        ]
        combined_text = "\n\n".join(parts)
        # 해당 페이지(이전, 현재, 다음)별 elementid와 페이지 리스트 생성
        prev_ids = df[df["페이지숫자"] == (current_page - 1)]["elementid"].tolist()
        curr_ids = df[df["페이지숫자"] == current_page]["elementid"].tolist()
        next_ids = df[df["페이지숫자"] == (current_page + 1)]["elementid"].tolist()
        all_ids = prev_ids + curr_ids + next_ids
        all_pages = ([current_page - 1] * len(prev_ids)) + ([current_page] * len(curr_ids)) + ([current_page + 1] * len(next_ids))
        meta_obj = {
            "elementid": all_ids,
            "category": row["data-category"],
            "filename": row["파일명"],
            "page": all_pages,
            "text": combined_text
        }
        metadata.append(meta_obj)
    return metadata

def get_cross_page_metadata(df):
    """
    metadata -3 생성 함수 (페이지 기반)  
    → 현재페이지 전체내용, 이전페이지 마지막 청크, 다음페이지 첫번째 청크를 라벨과 함께 결합하고,
       현재 페이지 그룹의 모든 elementid, 그리고 이전/다음 페이지의 해당 청크 정보를 리스트로 기록
       
    JSON 객체 형식:
    {
        "elementid": [int, ...],  # 현재 페이지의 모든 elementid + (이전페이지 마지막, 다음페이지 첫번째)
        "category": str,           # 현재 행의 data-category
        "filename": str,           # 파일명
        "page": [int, ...],        # 현재 페이지의 번호 리스트 + 이전, 다음 페이지 번호 (각각 1개씩)
        "text": str             # 라벨 포함 결합 텍스트
    }
    """
    df["elementid"] = df["elementid"].astype(int)
    df["페이지숫자"] = df["페이지숫자"].astype(int)
    page_groups = df.groupby("페이지숫자")
    page_dict = {}
    for page, group in page_groups:
        group_sorted = group.sort_values("elementid")
        full_text = "\n\n".join(group_sorted["내용"].tolist())
        first_chunk = group_sorted.iloc[0]["내용"]
        last_chunk = group_sorted.iloc[-1]["내용"]
        elem_ids = group_sorted["elementid"].tolist()
        page_dict[page] = {"full": full_text, "first": first_chunk, "last": last_chunk, "ids": elem_ids}
    metadata = []
    for _, row in df.iterrows():
        current_page = int(row["페이지숫자"])
        current_full = page_dict.get(current_page, {}).get("full", "")
        prev_last = page_dict.get(current_page - 1, {}).get("last", "")
        next_first = page_dict.get(current_page + 1, {}).get("first", "")
        parts = [
            "[[[[[[현재페이지 전체내용]", current_full,
            "[[[[[[이전페이지 마지막 청크]", prev_last,
            "[[[[[[다음페이지 첫번째 청크]", next_first
        ]
        combined_text = "\n\n".join(parts)
        current_ids = page_dict.get(current_page, {}).get("ids", [])
        prev_ids = df[df["페이지숫자"] == (current_page - 1)]["elementid"].tolist()
        next_ids = df[df["페이지숫자"] == (current_page + 1)]["elementid"].tolist()
        prev_last_id = [prev_ids[-1]] if prev_ids else []
        next_first_id = [next_ids[0]] if next_ids else []
        all_ids = current_ids + prev_last_id + next_first_id
        current_pages = [current_page] * len(current_ids)
        prev_page_list = [current_page - 1] if prev_ids else []
        next_page_list = [current_page + 1] if next_ids else []
        all_pages = current_pages + prev_page_list + next_page_list
        meta_obj = {
            "elementid": all_ids,
            "category": row["data-category"],
            "filename": row["파일명"],
            "page": all_pages,
            "text": combined_text
        }
        metadata.append(meta_obj)
    return metadata

def save_excel(content_list, metadata_list, output_path):
    # metadata_list의 각 JSON 객체를 줄바꿈이 포함된 문자열로 변환 (indent=4)
    metadata_json = [json.dumps(item, ensure_ascii=False, indent=4) for item in metadata_list]
    out_df = pd.DataFrame({
        "content": content_list,
        "metadata": metadata_json
    })
    out_df.to_excel(output_path, index=False)
    
    # openpyxl을 사용하여 엑셀 서식 적용
    wb = load_workbook(output_path)
    ws = wb.active
    ws.column_dimensions['A'].width = 80
    ws.column_dimensions['B'].width = 80
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    wb.save(output_path)

def construct_embedding_contents(base_folder):
    base_name = os.path.basename(os.path.normpath(base_folder))
    excel_path = os.path.join(base_folder, f"{base_name}.xlsx")
    if not os.path.exists(excel_path):
        print(f"❌ 엑셀 파일이 존재하지 않습니다: {excel_path}")
        return

    df = pd.read_excel(excel_path)
    # "이미지설명" 열이 있다면, 해당 행의 "내용"에 추가하여 현재 청크의 내용에 포함시킴
    if "이미지설명" in df.columns:
        df["내용"] = df.apply(
            lambda row: row["내용"] + ( "\n\n이미지설명: " + str(row["이미지설명"]) 
                                        if pd.notna(row["이미지설명"]) and str(row["이미지설명"]).strip() != "" 
                                        else "" ),
            axis=1
        )
    
    os.makedirs(os.path.join(base_folder, "before"), exist_ok=True)
    # content 생성
    content_map = {
        "chunk_only": df["내용"].tolist(),  # 원본 청크 (이미지설명 포함)
        "chunk_with_neighbors": extract_neighbors_by_elementid(df),
        "page_plus_chunk": extract_page_plus_chunk(df),
        "page_only": extract_page_only(df)
    }
    content_type_mapping = {
        "chunk_only": "1",
        "chunk_with_neighbors": "2",
        "page_plus_chunk": "3",
        "page_only": "4"
    }
    metadata_funcs = {
        "1": get_neighbor_metadata,
        "2": get_3page_metadata,
        "3": get_cross_page_metadata
    }
    # 각 content type별로 메타데이터 생성 (chunk_only, chunk_with_neighbors, page_plus_chunk: -1, -2, -3 / page_only: -2, -3)
    for content_name, content_list in content_map.items():
        if content_name == "page_only":
            valid_meta_ids = ["2", "3"]
        else:
            valid_meta_ids = ["1", "2", "3"]
        for meta_id in valid_meta_ids:
            metadata_list = metadata_funcs[meta_id](df)
            filename = f"{content_type_mapping[content_name]}-{meta_id}_{content_name}.xlsx"
            save_path = os.path.join(base_folder, "before")
            save_file_path = os.path.join(save_path, filename)
            save_excel(content_list, metadata_list, save_file_path)
    print("║ 📁 총 11개의 content|metadata 엑셀 파일이 생성되었습니다.")

    return save_path

if __name__ == "__main__":
    base_folder = "data/250331-13-24_모니터1~3p"
    construct_embedding_contents(base_folder)
